from tkinter import Tk , Label , Button , filedialog ,ttk ,constants as con , Frame , Text , messagebox as msg
import xlrd as xl
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning) 
import re

wb = None
filename = ""

def openExcelFile():
    global wb , filename
    filetypes = (
        ('Excel Files', '*.xlsx'),
        ('All files', '*.*')
    )
    filename = filedialog.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)
    
    fileSelected.config(text=filename)
    if filename != '':
        wb = load_workbook(filename)
        comboSheets.config(values=wb.sheetnames)
    else:
        fileSelected.config(text="Select file...")


def getColumnsBySheet(e):
    sheet = wb.get_sheet_by_name(comboSheets.get())
    list1 = []
    for cell in sheet[1]:
        list1.append(cell.value)
    comboColumns.config(values=list1)

def formatSpacing():
    sheet = comboSheets.get()
    columnName = comboColumns.get()
    if filename == "" or sheet == "" or columnName == "":
        msg.showinfo("Info" , "select file , sheet and column name to format")
        return
    sheet = wb.get_sheet_by_name(comboSheets.get())
    columnName = comboColumns.get()

    i = 0
    for column in sheet[1]:
        if(column.value == columnName):
            break
        i+=1

    j = 0       
    for row in sheet:
        
        newColString = ""
        splits = re.split(r"\s{2}",row[i].value)
        for each in splits:
            match = re.search(r"[A-Z0-9a-z]" , each)
            if(match!=None):
                newColString += re.sub(r"\A\s|\Z\s" , "" ,each) + "\n"
        row[i].value = newColString
        j+=1
    msg.showinfo("Success" , "Format Complete")
    

# def formatKeys():
#     global filename
#     sheet = wb.get_sheet_by_name(comboSheets.get())
#     columnName = comboColumns.get()

#     keys = keyValue.get(0.0 , con.END).split("\n")

#     i = 0
#     for column in sheet[1]:
#         if(column.value == columnName):
#             break
#         i+=1

#     j = 0       
#     for row in sheet:
#         if(j>1):
#             newColString = ""
#             splits = re.split(r"\s{2}",row[i].value)
#             for each in splits:
#                 match = re.search(r"[A-Z0-9a-z]" , each)
#                 if(match!=None):
#                     newColString += re.sub(r"\A\s|\Z\s" , "" ,each) + "\n"
#             k = 0
#             splits = newColString.split("\n")
#             s = {}
#         j+=1

def export():
    global wb , filename
    if wb!=None:
        savefilename = filedialog.asksaveasfile(initialfile = filename.split("/")[-1],defaultextension=".xlsx",filetypes=[("All Files","*.*"),("Excel files","*.xlsx")])
        if savefilename != None:
            wb.save(savefilename.name)
            wb = None
            filename = ""
            fileSelected.config(text="Select file...")
            comboColumns.config(state="normal")
            comboColumns.delete(0,con.END)
            comboColumns.config(state="readonly")
            comboColumns.config(values=[])
            comboSheets.config(state="normal")
            comboSheets.delete(0,con.END)
            comboSheets.config(values=[])
            comboSheets.config(state="readonly")
            msg.showinfo("Success" , "Export Complete")
    else:
        msg.showinfo("Info" , "Select a sheet and format before exporting")
    

root = Tk()
root.geometry("450x210")
root.title("Excel Formatter")
root.resizable(con.FALSE , con.FALSE)

Button(root, text = "Select File" , command=openExcelFile).place(x = 30,y = 50) 
Label(root, text = "Select Sheet : ").place(x = 30,y = 80)  
comboSheets = ttk.Combobox(
    state="readonly",
)
comboSheets.bind("<<ComboboxSelected>>", getColumnsBySheet)
Label(root, text = "Select Column : ").place(x = 30,y = 110)  
comboColumns = ttk.Combobox(
    state="readonly",
)

Button(root, text = "Format Spacing" , command=formatSpacing).place(x = 30,y = 150)  
#Label(root, text = "Enter Keys : ").place(x = 30,y = 180)  
#keyValue = Text(root , height = 10 , width = 30)
#Button(root, text = "Format Keys" , command=formatKeys).place(x = 30,y = 230)  

Button(root, text = "Export" , command=export).place(x = 200,y = 150) 

fileSelected =  Label(root, text = "Select file...")
fileSelected.place(x = 150,y = 50)  
comboSheets.place(x = 150,y = 80)
comboColumns.place(x = 150,y = 110)
#keyValue.place(x = 150 , y = 180)
root.mainloop()